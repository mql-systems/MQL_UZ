#!/usr/bin/env python3
# -*- coding: utf-8 -*-


# from bs4 import BeautifulSoup
# import csv
# from datetime import datetime
# import os
# import requests
# import json
# import matplotlib.pyplot as plt
# from openpyxl.styles import PatternFill
from openpyxl import Workbook
import pandas as pd


df = pd.read_csv(
    filepath_or_buffer='XAUUSD1440.csv',
    names={
        'Date': 'date',
        'Open': 'open',
        'High': 'high',
        'Low': 'low',
        'Close': 'close',
        'Volume': 'volume'
    }
)

# 2-qatordan boshlab o'qiydi
df = df.iloc[1:]

wb = Workbook()
ws = wb.active

# Shapka
ws['A1'] = "Kun"
ws['B1'] = "Hafta kuni"
ws['C1'] = "Sham uzunligi"
ws['D1'] = "Sham tanasi"
ws['E1'] = "Foiz"
ws['F1'] = "Sham soyasi"
ws['G1'] = "Foiz"
ws['H1'] = "Sham yuqori soyasi"
ws['I1'] = "Foiz"
ws['J1'] = "Sham pastgi soyasi"
ws['K1'] = "Foiz"
ws['L1'] = "Volume"
ws['M1'] = "Kun"
ws['N1'] = "Hafta kuni"
ws['O1'] = "Foyda"
ws['P1'] = "Prasadka"

f = []
y = 2
Date = df['Date']
Open = df['Open'].astype(float)
High = df['High'].astype(float)
Low = df['Low'].astype(float)
Close = df['Close'].astype(float)
Volume = df['Volume'].astype(int)
kun = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba']
text = ""

for i in range(2, len(Date)):

    foyda = 0
    prosadka = 0

    df_open_time = pd.to_datetime(Date[i])
    week_day = df_open_time.weekday()
    candle_length_1 = High[i - 1] - Low[i - 1]
    length_candle_body_1 = abs(Open[i - 1] - Close[i - 1])

    candle_buy_1 = Open[i - 1] < Close[i - 1]
    candle_sell_1 = Open[i - 1] > Close[i - 1]
    cl = candle_length_1 - length_candle_body_1
    upper_part = High[i - 1] - Open[i - 1] if Open[i - 1] > Close[i - 1] else High[i - 1] - Close[i - 1]
    lower_part = Close[i - 1] - Low[i - 1] if Open[i - 1] > Close[i - 1] else Open[i - 1] - Low[i - 1]

    if candle_buy_1:
        foyda = round(Open[i] - Low[i], 5)
        prosadka = round(High[i] - Open[i], 5)
    else:
        foyda = round(High[i] - Open[i], 5)
        prosadka = round(Open[i] - Low[i], 5)
    if foyda >= 0:
        ws['A' + str(y)] = df['Date'][i - 1]
        ws['B' + str(y)] = kun[pd.to_datetime(Date[i - 1]).weekday()]
        ws['C' + str(y)] = str(round(candle_length_1, 5)).replace('.', ',')
        ws['D' + str(y)] = str(round(length_candle_body_1, 5)).replace('.', ',')
        ws['E' + str(y)] = str(int(length_candle_body_1 / (candle_length_1 * 0.01)))
        ws['F' + str(y)] = str(round(cl, 5)).replace('.', ',')
        ws['G' + str(y)] = str(int(cl / (candle_length_1 * 0.01)))
        ws['H' + str(y)] = str(round(upper_part, 5)).replace('.', ',')
        ws['I' + str(y)] = str(int(upper_part / (candle_length_1 * 0.01)))
        ws['J' + str(y)] = str(round(lower_part, 5)).replace('.', ',')
        ws['K' + str(y)] = str(int(lower_part / (candle_length_1 * 0.01)))
        ws['L' + str(y)] = str(Volume[i - 1])
        ws['M' + str(y)] = df['Date'][i]
        ws['N' + str(y)] = kun[week_day]
        ws['O' + str(y)] = str(foyda).replace('.', ',')
        ws['P' + str(y)] = str(prosadka).replace('.', ',')
        y += 1

wb.save('XAUUSD1440.xlsx')
